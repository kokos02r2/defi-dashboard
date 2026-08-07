"""Загрузка выписки из Excel или CSV.

Жёсткого парсера здесь нет и быть не может: у каждого банка своя шапка, свой порядок
колонок, свой знак у расхода и своя запись даты. Поэтому импорт устроен в два шага —
сначала файл читается «как есть» и показывается предпросмотр с угаданными колонками,
потом человек подтверждает или исправляет сопоставление. Угадывание экономит время,
но не решает за пользователя.

Сопоставление запоминается на счёте: выписки одного банка приходят в одном формате,
и второй раз указывать колонки не нужно.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime

log = logging.getLogger(__name__)

SUPPORTED = (".xlsx", ".xlsm", ".csv", ".txt", ".tsv")

# Сколько строк показывать в предпросмотре. Больше не нужно: чтобы понять формат,
# хватает нескольких, а тысяча строк в браузере просто не читается.
PREVIEW_ROWS = 12
# Ограничение на размер файла: выписка за год — это десятки килобайт, а десяток
# мегабайт означает, что загрузили не то.
MAX_BYTES = 8 * 1024 * 1024

# Слова в шапке, по которым угадываются колонки. Русский, английский и испанский:
# карты в разных банках, и шапка бывает на любом из трёх.
HINTS = {
    "day": ("дата операции", "дата начала", "дата проводки", "дата", "date",
            "fecha operación", "fecha valor", "fecha", "f. valor", "f. operación",
            "buchung", "datum"),
    # Комиссия банка — отдельные деньги, а не часть суммы операции, поэтому у неё
    # своя колонка, а не подсказка для «суммы».
    "fee": ("комиссия", "сумма комиссии", "fee", "commission", "comisión"),
    # Состояние операции: отменённые и отклонённые лежат в выписке рядом с обычными
    "state": ("state", "статус", "состояние", "status", "estado"),
    "amount": ("сумма операции", "сумма в валюте счёта", "сумма", "amount", "importe",
               "importe eur", "betrag", "total"),
    "expense": ("расход", "списание", "debit", "cargo", "gasto", "salida", "debe"),
    "income": ("приход", "поступлени", "зачислен", "credit", "abono", "ingreso",
               "haber"),
    "note": ("описание", "назначение платежа", "назначение", "детали", "контрагент",
             "получатель", "место", "description", "concepto", "concepto ampliado",
             "descripción", "detalle", "merchant", "beneficiario", "text"),
    "currency": ("валюта операции", "валюта", "currency", "divisa", "moneda"),
}
# Значения колонки состояния, при которых денег не двигалось. Отменённую операцию
# банк оставляет в выписке навсегда — записать её значило бы навсегда завысить траты.
# «В обработке» сюда не входит: такая операция обычно доходит, а при следующей
# загрузке выписки она же распознается как повтор и второй раз не запишется.
BAD_STATE = ("отмен", "отклон", "не проведен", "cancel", "declin", "revert",
             "reject", "fail")
# Колонки, которые внешне похожи на сумму, но суммой операции не являются: остаток по
# счёту после операции угадывается как «сумма» и молча ломает весь импорт.
NOT_AMOUNT = ("остаток", "баланс", "saldo", "balance", "saldo posterior")

FIELDS = ("day", "amount", "expense", "income", "note", "currency", "fee", "state")


@dataclass
class Mapping:
    """Какая колонка что означает. Индексы — от нуля, None — колонка не используется."""
    header_row: int = 0
    day: int | None = None
    amount: int | None = None          # одна колонка со знаком
    expense: int | None = None         # либо две отдельные: расход...
    income: int | None = None          # ...и приход
    note: int | None = None
    currency: int | None = None
    fee: int | None = None             # комиссия сверх суммы операции
    state: int | None = None           # состояние: отменённые строки не записываются
    # В большинстве выписок расход записан отрицательным числом. Но не во всех:
    # некоторые банки отдают колонку «расход» положительными значениями.
    expense_negative: bool = True

    def to_dict(self) -> dict:
        return {f: getattr(self, f) for f in FIELDS} | {
            "header_row": self.header_row, "expense_negative": self.expense_negative}

    @classmethod
    def from_dict(cls, d: dict | None) -> "Mapping":
        d = d or {}
        m = cls()
        for f in FIELDS:
            v = d.get(f)
            setattr(m, f, int(v) if isinstance(v, int) or (isinstance(v, str) and v.isdigit()) else None)
        m.header_row = int(d.get("header_row") or 0)
        m.expense_negative = bool(d.get("expense_negative", True))
        return m

    @property
    def ok(self) -> bool:
        return self.day is not None and (
            self.amount is not None or self.expense is not None or self.income is not None)


@dataclass
class Row:
    """Разобранная строка файла. Пустой `error` означает, что её можно сохранять."""
    index: int
    day: date | None = None
    kind: str = "expense"
    amount: float = 0.0
    currency: str = ""
    note: str = ""
    error: str = ""
    # Пропуск — это не ошибка разбора: строка прочитана, но записывать её нечего
    # (операция отменена банком или в сумме ноль). Разные вещи и в счётчиках разные.
    skip: str = ""
    raw: list[str] = field(default_factory=list)


# --------------------------------------------------------------------------------------
# Чтение файла
# --------------------------------------------------------------------------------------

def read_table(filename: str, data: bytes) -> list[list]:
    """Файл -> таблица значений. Ячейки остаются как есть: дата из Excel приходит
    объектом datetime, и превращать её в строку, чтобы потом разбирать обратно, глупо."""
    if len(data) > MAX_BYTES:
        raise ValueError(f"файл больше {MAX_BYTES // (1024 * 1024)} МБ — это точно не выписка")
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        return _read_xlsx(data)
    if name.endswith((".csv", ".txt", ".tsv")):
        return _read_csv(data)
    if name.endswith(".xls"):
        raise ValueError("старый формат .xls не читается — сохраните файл как .xlsx или .csv")
    raise ValueError("поддерживаются .xlsx и .csv")


def _read_xlsx(data: bytes) -> list[list]:
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover — зависимость есть в requirements
        raise ValueError("чтение Excel недоступно: не установлен openpyxl") from None
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    return _trim(rows)


def _read_csv(data: bytes) -> list[list]:
    text = None
    for enc in ("utf-8-sig", "cp1251", "latin-1"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("не удалось определить кодировку файла")
    # Разделитель определяем по первым строкам: точка с запятой в европейских
    # выписках встречается чаще запятой, потому что запятая занята дробной частью.
    head = "\n".join(text.splitlines()[:20])
    delim = max((";", ",", "\t", "|"), key=head.count)
    if head.count(delim) == 0:
        delim = ";"
    return _trim([list(r) for r in csv.reader(io.StringIO(text), delimiter=delim)])


def _trim(rows: list[list]) -> list[list]:
    """Убирает полностью пустые строки по краям и выравнивает длину."""
    def empty(r: list) -> bool:
        return all(c is None or str(c).strip() == "" for c in r)
    while rows and empty(rows[0]):
        rows.pop(0)
    while rows and empty(rows[-1]):
        rows.pop()
    width = max((len(r) for r in rows), default=0)
    return [list(r) + [None] * (width - len(r)) for r in rows]


# --------------------------------------------------------------------------------------
# Разбор значений
# --------------------------------------------------------------------------------------

_NUM = re.compile(r"-?[\d\s .,]+")


def parse_number(v) -> float | None:
    """Сумма из выписки: «1 234,56», «-1.234,56», «(45.20)», «1,234.56», «45,20 EUR».

    Десятичный разделитель определяется по последнему из знаков — это надёжнее любого
    предположения о локали: в «1.234,56» последняя запятая, в «1,234.56» последняя
    точка. Скобки вокруг числа означают минус — так расход пишет часть банков.
    """
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()").replace(" ", " ")
    m = _NUM.search(s)
    if not m:
        return None
    s = m.group(0).strip().replace(" ", "")
    if not s or s in ("-", ".", ","):
        return None
    last_dot, last_comma = s.rfind("."), s.rfind(",")
    if last_dot >= 0 and last_comma >= 0:
        dec = "." if last_dot > last_comma else ","
        s = s.replace("," if dec == "." else ".", "").replace(dec, ".")
    elif last_dot >= 0 or last_comma >= 0:
        # Один разделитель, и он неоднозначен: «1.234» — это и тысяча двести
        # тридцать четыре (европейская запись), и одна целая с копейками. Решает
        # число знаков после него: в деньгах дробная часть — две цифры, а ровно
        # три означают разделитель тысяч. Правило одинаково для точки и запятой,
        # потому что банки пишут и так, и так.
        sep = "." if last_dot >= 0 else ","
        pos = max(last_dot, last_comma)
        tail = s[pos + 1:]
        s = s.replace(sep, "") if len(tail) == 3 and pos > 0 and tail.isdigit() \
            else s.replace(sep, ".")
    try:
        val = float(s)
    except ValueError:
        return None
    return -val if neg else val


_DATE_FORMATS = ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y",
                 "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y", "%d %m %Y")


def parse_day(v) -> date | None:
    """Дата операции. День-первым, а не месяц-первым: и русские, и европейские выписки
    пишут 05.03 как пятое марта. Формат «месяц первым» пробуется последним, когда
    первое число заведомо больше двенадцати."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    s = s.split(" ")[0] if " " in s and len(s.split(" ")[0]) >= 6 else s
    s = s.replace("T", " ").split(" ")[0]
    for f in _DATE_FORMATS:
        try:
            return datetime.strptime(s, f).date()
        except ValueError:
            continue
    return None


_CUR_SYMBOLS = {"€": "EUR", "$": "USD", "₽": "RUB", "£": "GBP", "₺": "TRY", "₸": "KZT"}


def parse_currency(v, default: str) -> str:
    raw = str(v or "")
    for sym, code in _CUR_SYMBOLS.items():
        if sym in raw:
            return code
    s = re.sub(r"[^A-Za-zА-Яа-я]", "", raw).upper()[:3]
    if s in ("РУБ", "RUR"):
        return "RUB"
    return s if len(s) == 3 and s.isascii() else default


# --------------------------------------------------------------------------------------
# Угадывание колонок
# --------------------------------------------------------------------------------------

def guess(rows: list[list]) -> Mapping:
    """Ищет строку шапки и сопоставляет колонки по её словам.

    Если шапки нет вовсе (выгрузка без заголовков), колонки определяются по содержимому
    первых строк: где даты — там дата, где числа — там сумма.
    """
    m = Mapping()
    best_score, best_row = 0, None
    for i, row in enumerate(rows[:15]):
        score = sum(1 for c in row if _match_field(c))
        if score > best_score:
            best_score, best_row = score, i
    if best_row is not None and best_score >= 2:
        m.header_row = best_row
        for j, cell in enumerate(rows[best_row]):
            f = _match_field(cell)
            if f and getattr(m, f) is None:
                setattr(m, f, j)
    else:
        m.header_row = -1        # шапки нет, данные начинаются с первой строки
        _guess_by_content(rows, m)

    # Пара «расход/приход» и одиночная колонка суммы исключают друг друга
    if m.expense is not None or m.income is not None:
        m.amount = None
    if m.day is None or (m.amount is None and m.expense is None and m.income is None):
        _guess_by_content(rows[max(m.header_row + 1, 0):], m)
    return m


def _match_field(cell) -> str | None:
    s = re.sub(r"\s+", " ", str(cell or "").strip().lower())
    if not s or len(s) > 60:
        return None
    if any(bad in s for bad in NOT_AMOUNT):
        return None
    for f in ("day", "state", "fee", "expense", "income", "amount", "currency", "note"):
        for word in HINTS[f]:
            if s == word or s.startswith(word) or word in s:
                return f
    return None


def _guess_by_content(rows: list[list], m: Mapping) -> None:
    """Резервное угадывание: по типу значений в первых строках данных."""
    sample = [r for r in rows[:20] if any(c is not None and str(c).strip() for c in r)]
    if not sample:
        return
    width = max(len(r) for r in sample)
    for j in range(width):
        col = [r[j] if j < len(r) else None for r in sample]
        dates = sum(1 for c in col if parse_day(c) is not None)
        nums = sum(1 for c in col if parse_number(c) is not None and parse_day(c) is None)
        texts = sum(1 for c in col if isinstance(c, str) and len(c.strip()) > 6
                    and parse_number(c) is None)
        half = max(len(col) // 2, 1)
        if m.day is None and dates >= half:
            m.day = j
        elif m.amount is None and m.expense is None and nums >= half:
            m.amount = j
        elif m.note is None and texts >= half:
            m.note = j


# --------------------------------------------------------------------------------------
# Предпросмотр и разбор
# --------------------------------------------------------------------------------------

def data_rows(rows: list[list], m: Mapping) -> list[list]:
    return rows[m.header_row + 1:] if m.header_row >= 0 else rows


def header_labels(rows: list[list], m: Mapping) -> list[str]:
    """Названия колонок для формы сопоставления. Без шапки — «Колонка 1, 2, 3…»."""
    width = max((len(r) for r in rows), default=0)
    if m.header_row >= 0 and m.header_row < len(rows):
        head = rows[m.header_row]
        return [(str(head[j]).strip() if j < len(head) and head[j] is not None
                 and str(head[j]).strip() else f"Колонка {j + 1}") for j in range(width)]
    return [f"Колонка {j + 1}" for j in range(width)]


def parse_rows(rows: list[list], m: Mapping, default_currency: str,
               limit: int | None = None, expenses_only: bool = True,
               not_before: date | None = None) -> list[Row]:
    """Разбирает строки данных по сопоставлению. Ошибочные не выбрасываются, а
    возвращаются с заполненным error: человек должен видеть, что не прочиталось.

    Из выписки берутся только расходы. Приход в файле — это в основном перекладывание
    собственных денег: возвраты, переводы с других своих счетов, обмен валюты. Считать
    это доходом значит завысить и доходы, и сбережения; настоящие доходы заводятся
    руками, их за месяц единицы. Строки прихода не выбрасываются молча — они
    остаются в списке с пометкой, чтобы в предпросмотре было видно, что с ними стало.

    not_before отсекает старые годы. Банк отдаёт выписку за всю историю счёта, поэтому
    без отсечки удалённые вручную старые операции возвращались бы при следующей же
    загрузке того же файла.
    """
    out: list[Row] = []
    for i, raw in enumerate(data_rows(rows, m)):
        if limit is not None and len(out) >= limit:
            break
        if all(c is None or str(c).strip() == "" for c in raw):
            continue
        r = _parse_one(i, raw, m, default_currency)
        if expenses_only and not r.error and not r.skip and r.kind == "income":
            r.skip = "приход — в выписке не учитывается"
        if not_before is not None and not r.error and not r.skip \
                and r.day is not None and r.day < not_before:
            r.skip = f"раньше {not_before.strftime('%d.%m.%Y')} — не загружаем"
        out.append(r)
    return out


def _cell(raw: list, idx: int | None):
    return raw[idx] if idx is not None and idx < len(raw) else None


def _parse_one(i: int, raw: list, m: Mapping, default_currency: str) -> Row:
    r = Row(index=i, raw=[("" if c is None else str(c)) for c in raw])
    r.note = str(_cell(raw, m.note) or "").strip()[:300]
    r.currency = parse_currency(_cell(raw, m.currency), default_currency)

    if m.state is not None:
        state = str(_cell(raw, m.state) or "").strip()
        if state and any(bad in state.lower() for bad in BAD_STATE):
            r.skip = f"операция не прошла: {state.lower()}"
            return r

    r.day = parse_day(_cell(raw, m.day))
    if r.day is None:
        r.error = "не разобрана дата"
        return r
    if r.day > date.today():
        r.error = "дата в будущем"
        return r

    fee = parse_number(_cell(raw, m.fee)) if m.fee is not None else None
    if m.amount is not None:
        v = parse_number(_cell(raw, m.amount))
        if v is None and fee is None:
            r.error = "не разобрана сумма"
            return r
        # Приводим к виду «минус — это расход», чтобы дальше знак значил одно и то же
        # независимо от того, каким знаком банк помечает списание.
        total = (v or 0.0) if m.expense_negative else -(v or 0.0)
    else:
        exp = parse_number(_cell(raw, m.expense)) if m.expense is not None else None
        inc = parse_number(_cell(raw, m.income)) if m.income is not None else None
        if exp is None and inc is None and fee is None:
            r.error = "не разобрана сумма"
            return r
        # В паре колонок знак не несёт смысла: «расход» — это расход, чем бы он ни
        # был записан. Некоторые банки пишут там минус, некоторые нет.
        total = (inc or 0.0) - abs(exp or 0.0)

    # Комиссия списывается сверх суммы операции, поэтому вычитается всегда — и когда
    # своя строка целиком состоит из неё (сумма ноль), и когда банк вернул комиссию
    # обратно (она приходит с минусом и тогда прибавляется).
    total -= (fee or 0.0)
    if round(total, 2) == 0:
        r.skip = "нулевая сумма"
        return r
    r.kind = "expense" if total < 0 else "income"
    r.amount = abs(round(total, 2))
    return r
